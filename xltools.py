import json
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook
from datetime import datetime

def xl_to_json(xlfilepath, xlsheetname, outputname=None): 
    """
    Convert to json from xlsheet
    """
    wb = load_workbook(xlfilepath)
    sheet = wb['Risk Data']
    # sheet.auto_filter.ref = "A:N"
    # sheet.auto_filter.add_filter_column(0, ["2021-01-01 00:00:00"], blank=False)

    def max_row(sheet): 
        """
        Sheet (Openpyxl object) --> Get the maximum number of occupied cells in a XL worksheet 
        """
        count = 0
        for row in sheet:
            if not all([cell.value == None for cell in row]):
                count += 1 
        return count
    
#     def max_column(sheet): 
#         """
#         Sheet (Openpyxl object) --> Get the maximum number of occupied cells in a XL worksheet
#         """   
#         count = 0
#         for column in sheet:
#             if not all([cell.value == None for cell in column]):
#                 count += 1 
#         return count        

    def get_col_header_names(sheet, number_of_col): 
        """
        Sheet (Openpyxl object) --> Get the columns names in a list
        """
        col_names = []
        for k in range(1, number_of_col+1): 
                col_names.append(sheet.cell(row=1, column=k).value) 
        return col_names       
    
    #Initial list
    dict_list = []
    col_names = get_col_header_names(sheet, number_of_col=14)
    
    for row in islice(sheet.values, 1, max_row(sheet)): 
        dict_items = OrderedDict([(col_names[i], row[i]) for i in range(14)])
        dict_list.append(dict_items)
    
    # Serialize the list of dicts to JSON
    j = json.dumps(dict_list, indent=4, sort_keys=False, default=str)
    # Write to file
    with open(outputname + '_' + str(datetime.now().year) + str(datetime.now().month) + '.json', 'w') as f:
        f.write(j)

        
def merge_json_files(): 
    pass